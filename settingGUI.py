from ruamel.yaml import YAML
from tkinter import Tk, filedialog, messagebox, StringVar, Label, Button, Toplevel, Listbox, MULTIPLE
from tkinter.ttk import Combobox
import os
from openpyxl import load_workbook

# 定義設定檔路徑
DEFAULT_SETTING_YAML = "./setting.yaml"

# 初始化 YAML 物件
yaml = YAML()
yaml.preserve_quotes = True

def select_and_update_project_aim(setting_data):
    """
    使用 GUI 讓使用者選擇目前執行計畫 ('產學合作' 或 '研究計畫') 並更新設定資料。
    :param setting_data: 讀取的設定資料字典
    """
    root = Tk()
    root.title("計畫目標選擇")
    root.geometry("300x150")
    
    valid_project_aim_options = ["產學合作", "研究計畫"]

    selected_aim = StringVar()
    selected_aim.set(valid_project_aim_options[0])  # 預設值

    label = Label(root, text="請選擇目前執行計畫:")
    label.pack(pady=10)

    combobox = Combobox(root, textvariable=selected_aim, values=valid_project_aim_options, state="readonly")
    combobox.pack(pady=10)

    def on_select():
        current_project_aim = selected_aim.get()
        setting_data['SOURCE']['field']['目前執行計畫'] = current_project_aim
        messagebox.showinfo("成功", f"目前執行計劃目標已經設定為: {current_project_aim}")
        root.destroy()  # 正確關閉視窗

    button = Button(root, text="確認", command=on_select)
    button.pack(pady=10)
    
    root.mainloop()

def select_the_file_update_project_name(setting_data):
    """
    使用 GUI 讓使用者選擇一個檔案，並根據目前執行計畫更新對應的檔案名稱和副檔名。
    限制檔案選擇在 ./data 資料夾下的任何子資料夾。
    :param setting_data: 讀取的設定資料字典
    """
    root = Tk()
    root.title("選擇資料名稱")
    root.geometry("300x150")
    
    current_aim = setting_data['SOURCE']['field']['目前執行計畫']

    if current_aim == "研究計畫": 
        base_directory = os.path.abspath("./data/research_proj/")
    elif current_aim == "產學合作": 
        base_directory = os.path.abspath("./data/industry_coop/")

    file_window = Toplevel(root)
    file_window.title("選擇檔案")
    file_window.geometry("400x200")

    def open_file_dialog():
        file_path = filedialog.askopenfilename(initialdir=base_directory, title=f"選擇 {current_aim} 的初始資料檔案")
        if file_path:
            file_path = os.path.abspath(file_path)
            if file_path.startswith(base_directory):
                file_name = os.path.basename(file_path)
                if current_aim == '研究計畫':
                    setting_data['SOURCE']['data']['research_proj']['研究計畫申請名冊'] = file_name
                elif current_aim == '產學合作':
                    setting_data['SOURCE']['data']['industry_coop']['產學合作申請名冊'] = file_name
                messagebox.showinfo("成功", f"已更新檔案名稱為: {file_name}")
                file_window.destroy()  # 正確關閉視窗
                root.destroy()  # 關閉主視窗
                # 呼叫下一步驟的函數來選擇 Sheet
                select_sheet_from_excel(file_path, setting_data)
            else:
                messagebox.showerror("錯誤", "選擇的檔案不在允許的目錄內。")
        else:
            messagebox.showerror("錯誤", "未選擇任何檔案。")

    label = Label(file_window, text=f"請選擇 {current_aim} 的初始資料檔案")
    label.pack(pady=20)

    select_button = Button(file_window, text="選擇檔案", command=open_file_dialog)
    select_button.pack(pady=20)

    root.withdraw()
    file_window.mainloop()

def select_sheet_from_excel(file_path, setting_data):
    """
    讀取 Excel 檔案中的 Sheet，並讓使用者選擇要使用的 Sheet 名稱。
    更新設定資料的 '計畫SHEET' 欄位為 LIST。
    :param file_path: Excel 檔案的路徑
    :param setting_data: 讀取的設定資料字典
    """
    root = Tk()
    root.title("選擇計畫 SHEET (所有 SHEET 欄位記得統一)")
    root.geometry("300x300")

    # 讀取 Excel 檔案中的 Sheets
    workbook = load_workbook(file_path, read_only=True)
    sheets = workbook.sheetnames

    # 使用 Listbox 進行多選
    listbox = Listbox(root, selectmode=MULTIPLE)
    for sheet in sheets:
        listbox.insert('end', sheet)
    listbox.pack(pady=20)

    def on_select():
        selected_indices = listbox.curselection()
        selected_sheets = [sheets[i] for i in selected_indices]
        if selected_sheets:
            setting_data['SOURCE']['field']['計畫SHEET'] = selected_sheets
            messagebox.showinfo("成功", f"計畫 SHEET 已更新為: {', '.join(selected_sheets)}")
            root.destroy()  # 正確關閉視窗
        else:
            messagebox.showerror("錯誤", "至少選擇一個 SHEET。")

    button = Button(root, text="確認", command=on_select)
    button.pack(pady=10)

    root.mainloop()
    
def confirm_and_update_project_name_column(file_path, sheet_name, setting_data):
    """
    確認 Excel Sheet 中的欄位，並讓使用者選擇計畫相關的欄位。
    「計畫名稱」和「中文關鍵字」不能為空。
    其他欄位（如計畫相關其他欄位、申請機構、主持人、共同主持人、共同機構）可以為空。
    :param file_path: Excel 檔案的路徑
    :param sheet_name: Excel 中選擇的 Sheet 名稱
    :param setting_data: 讀取的設定資料字典
    """
    root = Tk()
    root.title("確認計畫相關欄位，如果必填沒有請去補齊!")
    root.geometry("500x800")

    # 讀取指定 Sheet 的欄位
    print(file_path, sheet_name)
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook[sheet_name]
    
    # 確保從第一列取得欄位名稱
    columns = [cell for cell in next(sheet.iter_rows(max_row=1, values_only=True))]

    # 設定選擇框的預設值
    selected_project_name_column = StringVar()
    selected_keyword_column = StringVar()
    selected_abstract_column = StringVar()
    selected_institution_column = StringVar()
    selected_lead_researcher_column = StringVar()

    # 設定「計畫名稱」和「中文關鍵字」的初始值
    current_project_name_column = setting_data['SOURCE']['field'].get('計畫名稱', '')
    current_keyword_column = setting_data['SOURCE']['field'].get('中文關鍵字', '')
    current_abstract_column = setting_data['SOURCE']['field'].get('計劃摘要', '')

    selected_project_name_column.set(current_project_name_column if current_project_name_column in columns else "")
    selected_keyword_column.set(current_keyword_column if current_keyword_column in columns else "")
    selected_abstract_column.set(current_abstract_column if current_abstract_column in columns else "")
    
    # 其他欄位可以為空，初始為空
    selected_institution_column.set("")
    selected_lead_researcher_column.set("")

    # 多選的欄位設置
    selected_co_lead_researchers_columns = []
    selected_co_institutions_columns = []
    selected_other_related_columns = []

    # 顯示計畫名稱選擇的 GUI
    project_name_label = Label(root, text="請選擇屬於計畫名稱的欄位:")
    project_name_label.pack(pady=5)

    project_name_combobox = Combobox(root, textvariable=selected_project_name_column, values=[""] + columns, state="readonly")
    project_name_combobox.pack(pady=5)

    # 顯示中文關鍵字選擇的 GUI
    keyword_label = Label(root, text="請選擇屬於中文關鍵字的欄位:")
    keyword_label.pack(pady=5)

    keyword_combobox = Combobox(root, textvariable=selected_keyword_column, values=[""] + columns, state="readonly")
    keyword_combobox.pack(pady=5)
    
    # 顯示中文關鍵字選擇的 GUI
    abstract_label = Label(root, text="請選擇屬於計劃摘要的欄位:")
    abstract_label.pack(pady=5)

    abstract_combobox = Combobox(root, textvariable=selected_abstract_column, values=[""] + columns, state="readonly")
    abstract_combobox.pack(pady=5)

    # 顯示申請機構選擇的 GUI
    institution_label = Label(root, text="請選擇屬於申請機構(學校)的欄位:")
    institution_label.pack(pady=5)

    institution_combobox = Combobox(root, textvariable=selected_institution_column, values=[""] + columns, state="readonly")
    institution_combobox.pack(pady=5)

    # 顯示主持人選擇的 GUI
    lead_researcher_label = Label(root, text="請選擇屬於(計畫)主持人的欄位:")
    lead_researcher_label.pack(pady=5)

    lead_researcher_combobox = Combobox(root, textvariable=selected_lead_researcher_column, values=[""] + columns, state="readonly")
    lead_researcher_combobox.pack(pady=5)

    # 顯示計畫相關其他欄位的多選 GUI
    other_related_fields_label = Label(root, text="請選擇計畫相關其他欄位 (可複選):")
    other_related_fields_label.pack(pady=5)

    other_related_fields_listbox = Listbox(root, selectmode=MULTIPLE, height=5, exportselection=False)
    for column in columns:
        other_related_fields_listbox.insert('end', column)
    other_related_fields_listbox.pack(pady=5)

    # 顯示共同主持人選擇的 GUI
    co_lead_researchers_label = Label(root, text="請選擇共同(計畫)主持人的欄位 (可複選):")
    co_lead_researchers_label.pack(pady=5)

    co_lead_researchers_listbox = Listbox(root, selectmode=MULTIPLE, height=5, exportselection=False)
    for column in columns:
        co_lead_researchers_listbox.insert('end', column)
    co_lead_researchers_listbox.pack(pady=5)

    # 顯示共同機構選擇的 GUI
    co_institutions_label = Label(root, text="請選擇共同機構(學校)的欄位 (可複選):")
    co_institutions_label.pack(pady=5)

    co_institutions_listbox = Listbox(root, selectmode=MULTIPLE, height=5, exportselection=False)
    for column in columns:
        co_institutions_listbox.insert('end', column)
    co_institutions_listbox.pack(pady=5)

    def on_select():
        # 確認「計畫名稱」和「中文關鍵字」是否選擇
        project_name_column = selected_project_name_column.get()
        keyword_column = selected_keyword_column.get()
        abstract_column = selected_abstract_column.get()

        # 單選的欄位
        institution_column = selected_institution_column.get()
        lead_researcher_column = selected_lead_researcher_column.get()

        # 多選的欄位
        selected_indices_other_related = other_related_fields_listbox.curselection()
        selected_other_related = [columns[i] for i in selected_indices_other_related]

        selected_indices_co_lead_researchers = co_lead_researchers_listbox.curselection()
        selected_co_lead_researchers = [columns[i] for i in selected_indices_co_lead_researchers]

        selected_indices_co_institutions = co_institutions_listbox.curselection()
        selected_co_institutions = [columns[i] for i in selected_indices_co_institutions]

        # 更新設定資料
        setting_data['SOURCE']['field']['計畫名稱'] = project_name_column
        setting_data['SOURCE']['field']['中文關鍵字'] = keyword_column
        setting_data['SOURCE']['field']['計劃摘要'] = abstract_column
        setting_data['SOURCE']['field']['申請機構欄位名稱'] = institution_column
        setting_data['SOURCE']['field']['申請主持人欄位名稱'] = lead_researcher_column
        setting_data['SOURCE']['field']['計畫相關其他欄位'] = selected_other_related
        setting_data['SOURCE']['field']['申請共同主持人'] = selected_co_lead_researchers
        setting_data['SOURCE']['field']['申請共同機構欄位名稱'] = selected_co_institutions

        messagebox.showinfo(
            "成功",
            f"計畫名稱欄位已設定為: {project_name_column}\n"
            f"中文關鍵字欄位已設定為: {keyword_column}\n"
            f"計劃摘要欄位已設定為: {abstract_column}\n"
            f"申請機構欄位已設定為: {institution_column}\n"
            f"申請主持人欄位已設定為: {lead_researcher_column}\n"
            f"計畫相關其他欄位已設定為: {', '.join(selected_other_related) if selected_other_related else '無'}\n"
            f"申請共同主持人已設定為: {', '.join(selected_co_lead_researchers) if selected_co_lead_researchers else '無'}\n"
            f"申請共同機構已設定為: {', '.join(selected_co_institutions) if selected_co_institutions else '無'}"
        )
        root.destroy()

    button = Button(root, text="確認", command=on_select)
    button.pack(pady=10)
    
    root.mainloop()


try:
    with open(DEFAULT_SETTING_YAML, 'r', encoding='utf-8') as file:
        setting_data = yaml.load(file)

    select_and_update_project_aim(setting_data)
    with open(DEFAULT_SETTING_YAML, 'w', encoding='utf-8') as file:
        yaml.dump(setting_data, file)
        
    select_the_file_update_project_name(setting_data)
    with open(DEFAULT_SETTING_YAML, 'w', encoding='utf-8') as file:
        yaml.dump(setting_data, file)

    # 根據目前執行計畫來決定檔案路徑
    current_aim = setting_data['SOURCE']['field']['目前執行計畫']
    if current_aim == "研究計畫":
        file_path = os.path.join(os.path.abspath("./data/research_proj/"), setting_data['SOURCE']['data']['research_proj']['研究計畫申請名冊'])
    elif current_aim == "產學合作":
        file_path = os.path.join(os.path.abspath("./data/industry_coop/"), setting_data['SOURCE']['data']['industry_coop']['產學合作申請名冊'])

    # 取得選擇的計畫 SHEET 名稱
    sheet_name = setting_data['SOURCE']['field']['計畫SHEET']
    confirm_and_update_project_name_column(file_path, sheet_name[0], setting_data)

    with open(DEFAULT_SETTING_YAML, 'w', encoding='utf-8') as file:
        yaml.dump(setting_data, file)

    messagebox.showinfo("完成", "設定已成功更新並存回檔案。")

except Exception as e:
    messagebox.showerror("錯誤", f"發生錯誤，無法更新設定檔。\n{e}")
finally:
    pass